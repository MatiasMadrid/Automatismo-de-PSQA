import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class RadioRiskApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Evaluación de Riesgo en Radioterapia")
        self.ancho_fijo = 480
        self.alto_fijo = 600
        self.root.geometry(f"{self.ancho_fijo}x{self.alto_fijo}")

        self.datos_paciente = {}
        self.entries = {}
        self.intento_actual = 1
        self.historial_intentos = {}  # Para guardar qué se usó en cada intento para el Excel

        self.archivo_config = "config_ruta.txt"
        self.ruta_informe = self.cargar_ruta_persistente()

        self.archivo_umbrales = "umbrales.txt"
        # Umbrales por defecto
        self.u_mcs = 0.5
        self.u_sas = 0.5
        self.u_fractions = 3
        self.cargar_umbrales()

        self.create_main_menu()

    # --- PERSISTENCIA DE UMBRALES ---
    def cargar_umbrales(self):
        if os.path.exists(self.archivo_umbrales):
            try:
                with open(self.archivo_umbrales, "r") as f:
                    lineas = f.readlines()
                    self.u_mcs = float(lineas[0].strip())
                    self.u_sas = float(lineas[1].strip())
                    self.u_fractions = int(lineas[2].strip())
            except:
                pass  # Si falla, usa los valores por defecto

    def guardar_umbrales_archivo(self, mcs, sas, frac):
        with open(self.archivo_umbrales, "w") as f:
            f.write(f"{mcs}\n{sas}\n{frac}")
        self.u_mcs, self.u_sas, self.u_fractions = mcs, sas, frac

    def validar_y_guardar_umbrales(self):
        try:
            # Sanitización: reemplazar coma por punto
            mcs = float(self.tmp_mcs.get().replace(',', '.'))
            sas = float(self.tmp_sas.get().replace(',', '.'))
            frac = int(self.tmp_frac.get())

            self.guardar_umbrales_archivo(mcs, sas, frac)
            messagebox.showinfo("Éxito", "Umbrales actualizados correctamente.")
            self.create_config_menu()
        except ValueError:
            messagebox.showerror("Error", "Por favor, ingrese solo números válidos.")

    def cargar_ruta_persistente(self):
        """Busca si existe una ruta guardada de una sesión anterior"""
        if os.path.exists(self.archivo_config):
            with open(self.archivo_config, "r") as f:
                ruta = f.read().strip()
                if os.path.exists(ruta):
                    return ruta
        return None

    def guardar_ruta_persistente(self, ruta):
        """Guarda la ruta en un archivo de texto para la próxima vez"""
        with open(self.archivo_config, "w") as f:
            f.write(ruta)
        self.ruta_informe = ruta

        # --- MENÚS DE NAVEGACIÓN ---
    def create_main_menu(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        menu_container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        menu_container.place(relx=0.5, rely=0.5, anchor="center")
        menu_container.pack_propagate(False)

        tk.Label(menu_container, text="Menú Principal", font=("Arial", 18, "bold")).pack(pady=40)

        tk.Button(menu_container, text="Cargar Paciente", width=25, height=2, bg="#E1E1E1",
                  command=self.cargar_archivo, font=("Arial", 10)).pack(pady=10)

        tk.Button(menu_container, text="Configuración", width=25, height=2, bg="#E1E1E1",
                  command=self.create_config_menu, font=("Arial", 10)).pack(pady=10)

        tk.Label(menu_container,
                 text=f"Ruta actual: {os.path.basename(self.ruta_informe) if self.ruta_informe else 'No configurada'}",
                 font=("Arial", 8, "italic"), fg="gray").pack(side="bottom", pady=20)

    def create_config_menu(self):
        for widget in self.root.winfo_children(): widget.destroy()
        c = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo);
        c.place(relx=0.5, rely=0.5, anchor="center");
        c.pack_propagate(False)

        tk.Label(c, text="Panel de Configuración", font=("Arial", 16, "bold")).pack(pady=20)

        tk.Button(c, text="Seleccionar Registro Existente", width=30, command=self.seleccionar_registro_existente).pack(pady=5)
        tk.Button(c, text="Crear Nuevo Registro", width=30, command=self.crear_nuevo_registro).pack(pady=5)
        tk.Button(c, text="Configurar Umbrales", width=30, command=self.create_thresholds_menu).pack(pady=5)
        tk.Button(c, text="Configurar Costos", width=30,command=self.abrir_excel_costos).pack(pady=5)
        tk.Button(c, text="Volver al Menú Principal", bg="#FFCCCB", command=self.create_main_menu).pack(side="bottom",pady=30)

    def create_thresholds_menu(self):
        for widget in self.root.winfo_children(): widget.destroy()
        container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        tk.Label(container, text="Configurar Umbrales de Complejidad",
                 font=("Arial", 14, "bold")).pack(pady=(30, 20))

        # Frame central que usará GRID para alineación de tabla
        frame_form = tk.Frame(container)
        frame_form.pack(expand=True)

        # Definir los widgets de entrada
        self.tmp_mcs = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_mcs.insert(0, str(self.u_mcs))

        self.tmp_sas = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_sas.insert(0, str(self.u_sas))

        self.tmp_frac = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_frac.insert(0, str(self.u_fractions))

        # Lista de campos para el bucle
        campos = [
            ("MCS Mínimo:", self.tmp_mcs),
            ("SAS Máximo:", self.tmp_sas),
            ("Fracciones Límite:", self.tmp_frac)
        ]

        # El truco del GRID: columna 0 para etiquetas, columna 1 para entradas
        for i, (texto, entry) in enumerate(campos):
            # Etiqueta: alineada a la derecha (sticky="e")
            tk.Label(frame_form, text=texto, font=("Arial", 10, "bold"),anchor="e", width=18).grid(row=i, column=0, pady=10, padx=10, sticky="e")

            # Cuadro de texto: alineado a la izquierda (sticky="w")
            entry.grid(row=i, column=1, pady=10, padx=10, sticky="w")

        # Área de botones
        btn_frame = tk.Frame(container)
        btn_frame.pack(pady=40)

        tk.Button(btn_frame, text="Guardar Cambios", bg="#4CAF50", fg="white",font=("Arial", 10, "bold"), width=20, height=2,command=self.validar_y_guardar_umbrales).pack(pady=5)

        tk.Button(btn_frame, text="Volver", font=("Arial", 10), width=15,command=self.create_config_menu).pack(pady=5)

    def seleccionar_registro_existente(self):
        """Busca un archivo Excel ya creado en el disco"""
        ruta = filedialog.askopenfilename(
            title="Seleccione el archivo de Registro Histórico existente",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if ruta:
            self.guardar_ruta_persistente(ruta)
            messagebox.showinfo("Éxito", f"Se ha vinculado el archivo:\n{ruta}")
            self.create_main_menu()

    def crear_nuevo_registro(self):
        """Define una ruta para un archivo que aún no existe"""
        nueva_ruta = filedialog.asksaveasfilename(
            title="Defina el nombre y ubicación del nuevo Registro",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Registro_Historico_2026.xlsx"
        )
        if nueva_ruta:
            self.guardar_ruta_persistente(nueva_ruta)
            messagebox.showinfo("Éxito",
                                f"Nuevo registro configurado en:\n{nueva_ruta}\n\nEl archivo se creará físicamente al exportar el primer paciente.")
            self.create_main_menu()

    def cambiar_ruta_manualmente(self):
        nueva_ruta = filedialog.asksaveasfilename(
            title="Seleccione o cree el archivo de Registro Histórico",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Registro_Historico_QA.xlsx"
        )
        if nueva_ruta:
            self.guardar_ruta_persistente(nueva_ruta)
            messagebox.showinfo("Configuración", f"Ruta actualizada correctamente:\n{nueva_ruta}")
            self.create_config_menu()  # Refrescar para mostrar cambios si fuera necesario

    def cargar_archivo(self):
        filepath = filedialog.askopenfilename(title="Seleccionar reporte", filetypes=[("Excel files", "*.xlsx *.xls")])
        if filepath:
            try:
                self.intento_actual = 1
                self.historial_intentos = {}
                self.extraer_datos(filepath)
                self.mostrar_detalles_paciente()
            except Exception as e:
                messagebox.showerror("Error", f"Error: {e}")

    def abrir_excel_costos(self):
        """Abre el archivo de costos con la aplicación predeterminada del sistema"""
        ruta_costos = "costos.xlsx"

        if os.path.exists(ruta_costos):
            try:
                # os.startfile funciona en Windows para abrir el archivo con Excel
                os.startfile(ruta_costos)
                messagebox.showinfo("Costos",
                                    "Abriendo el archivo de costos...\nRecuerde guardar los cambios en Excel antes de exportar un informe.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
        else:
            messagebox.showerror("Error", "No se encontró el archivo 'costos.xlsx' en la carpeta del proyecto.")

    def extraer_datos(self, path):
        df = pd.read_excel(path, header=None)

        def buscar_valor(etiqueta):
            for i, row in df.iterrows():
                for j, cell in enumerate(row):
                    if str(cell).strip() == etiqueta:
                        return str(df.iloc[i, j + 1]).strip()
            return "-"

        mcs_values, sas_values = [], []
        en_beam_metrics = False
        for i, row in df.iterrows():
            if str(row[0]).strip() == "BEAM METRICS":
                en_beam_metrics = True
                continue
            if en_beam_metrics:
                try:
                    metrica, valor_str = str(row[2]).strip(), str(row[3]).replace(',', '.')
                    if metrica == "MCS":
                        mcs_values.append(float(valor_str))
                    elif metrica == "SAS":
                        sas_values.append(float(valor_str))
                except:
                    continue

        self.datos_paciente = {
            "Plan": buscar_valor("PLAN NAME"),
            "Nombre": buscar_valor("PATIENT NAME"),
            "ID": buscar_valor("PATIENT ID"),
            "Sexo": buscar_valor("PATIENT SEX"),
            "Fractions": buscar_valor("FRACTIONS"),
            "MCS": buscar_valor("MCS"),
            "SAS": buscar_valor("SAS"),
            "PMU": buscar_valor("PMU"),
            "MCSmin": str(min(mcs_values)) if mcs_values else "-",
            "SASmax": str(max(sas_values)) if sas_values else "-"
        }

    def actualizar_checkbox_ca(self, *args):
        region = self.entries["Region"].get()
        regiones_con_ca = ["COLON/RECTO", "PULMON", "CERVIX/UTERO", "CYC"]
        self.entries["CA"].set(region in regiones_con_ca)

    def mostrar_detalles_paciente(self):
        for widget in self.root.winfo_children(): widget.destroy()

        container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        tk.Label(container, text="Información del Paciente", font=("Arial", 14, "bold")).pack(pady=10)
        frame_info = tk.LabelFrame(container, text=" Datos Extraídos ", padx=15, pady=10)
        frame_info.pack(padx=10, fill="both", expand=True)

        op_sexo, op_tecnica = ["M", "F", "-"], ["3D", "IMRT", "VMAT", "SRS", "SBRT", "FIF"]
        op_anatomica = ["MAMA", "COLON/RECTO", "PULMON", "PROSTATA", "CERVIX/UTERO", "ESOFAGO", "CYC", "PANCREAS",
                        "VEJIGA", "ENCEFALO/SNC", "MIEMBROS", "OTROS"]

        plan_name = self.datos_paciente.get("Plan", "").upper()
        palabras = plan_name.split()
        tecnica_def = next((t for t in op_tecnica if t in plan_name), "3D")
        region_def = palabras[2] if len(palabras) >= 3 and palabras[2] in op_anatomica else "OTROS"

        # Variables de control
        self.entries["Sexo"] = tk.StringVar(value=self.datos_paciente.get("Sexo", "-"))
        self.entries["Region"] = tk.StringVar(value=region_def)
        self.entries["Tecnica"] = tk.StringVar(value=tecnica_def)
        self.entries["CA"] = tk.BooleanVar()
        self.entries["PPed"] = tk.BooleanVar()
        self.entries["Region"].trace_add("write", self.actualizar_checkbox_ca)
        self.actualizar_checkbox_ca()

        # --- RESTAURACIÓN DE TODOS LOS CAMPOS ---
        campos = [
            ("ID Paciente", "ID"),
            ("Nombre", "Nombre"),
            ("Plan", "Plan"),
            ("MCS Prom.", "MCS"),
            ("SAS Prom.", "SAS"),
            ("PMU", "PMU"),
            ("MCS Mínimo", "MCSmin"),
            ("SAS Máximo", "SASmax"),
            ("Fracciones", "Fractions")
        ]
        for label, key in campos:
            row = tk.Frame(frame_info)
            row.pack(fill="x", pady=1)
            tk.Label(row, text=label, width=15, anchor="w", font=("Arial", 8)).pack(side="left")
            e = tk.Entry(row, font=("Arial", 8))
            e.insert(0, self.datos_paciente.get(key, "-"))
            e.config(state="readonly")
            e.pack(side="right", fill="x", expand=True)

        # Desplegables actualizados
        for lab, key, ops in [("Sexo", "Sexo", op_sexo), ("Región", "Region", op_anatomica),
                              ("Técnica", "Tecnica", op_tecnica)]:
            row = tk.Frame(frame_info);
            row.pack(fill="x", pady=2)
            tk.Label(row, text=lab, width=15, anchor="w", font=("Arial", 8, "bold")).pack(side="left")
            tk.OptionMenu(row, self.entries[key], *ops).pack(side="right", fill="x", expand=True)

        tk.Checkbutton(frame_info, text="Cambios Anatómicos", variable=self.entries["CA"], font=("Arial", 8)).pack(
            anchor="w")
        tk.Checkbutton(frame_info, text="Paciente Pediátrico", variable=self.entries["PPed"], font=("Arial", 8)).pack(
            anchor="w")

        tk.Button(container, text="Calcular Método QA", bg="#0078D7", fg="white", font=("Arial", 10, "bold"),
                  command=self.ejecutar_arbol_decision).pack(pady=10)
        tk.Button(container, text="Volver", command=self.create_main_menu).pack()

    # Empieza arbol
    def es_plan_complejo(self):
        """Determina la complejidad usando las variables configurables"""
        tecnica = self.entries["Tecnica"].get()
        if tecnica not in ["IMRT", "VMAT"]:
            return False

        try:
            mcs_paciente = float(self.datos_paciente.get("MCSmin", 1.0))
            sas_paciente = float(self.datos_paciente.get("SASmax", 0.0))
            frac_paciente = int(self.datos_paciente.get("Fractions", 1))

            # Lógica de comparación
            condicion_mcs = mcs_paciente < self.u_mcs
            condicion_sas = sas_paciente > self.u_sas
            condicion_frac = frac_paciente > self.u_fractions

            return condicion_mcs or condicion_sas or condicion_frac
        except:
            return False

    def obtener_paquete_qa(self):
        tecnica = self.entries["Tecnica"].get()
        ca_ped = self.entries["CA"].get() or self.entries["PPed"].get()
        complejo = self.es_plan_complejo()

        if tecnica in ["3D", "FIF"]:
            if self.intento_actual == 1:
                res = "Plancheck + Calculo independiente + LogFile"
                return res + " + Transit-EPID" if ca_ped else res
            return "Portal Dosimetry"

        elif tecnica in ["SRS", "SBRT"]:
            if self.intento_actual == 1:
                res = "Plancheck + Portal Dosimetry"
                return res + " + Transit-EPID" if ca_ped else res
            return "Stereophan + Gafchromic/CI"

        elif tecnica in ["IMRT", "VMAT"]:
            if not complejo:
                res = "Plancheck + Calculo independiente + LogFile"
                return res + " + Transit-EPID" if ca_ped else res
            else:
                if self.intento_actual == 1:
                    res = "Plancheck + Calculo independiente + LogFile + Portal Dosimetry"
                    return res + " + Transit-EPID" if ca_ped else res
                return "ArcCheck + 3DVH"
        return "Indefinido"

    def ejecutar_arbol_decision(self):
        for widget in self.root.winfo_children(): widget.destroy()
        container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        container.place(relx=0.5, rely=0.5, anchor="center");
        container.pack_propagate(False)

        paquete = self.obtener_paquete_qa()
        self.paquete_actual_str = paquete  # Guardar para el registro

        tk.Label(container, text="EVALUACIÓN QADS", font=("Arial", 14, "bold")).pack(pady=20)
        tk.Label(container, text=f"Intento N° {self.intento_actual}", font=("Arial", 10, "italic")).pack()

        lbl_paquete = tk.Label(container, text=paquete, font=("Arial", 11, "bold"), fg="#004080", wraplength=400)
        lbl_paquete.pack(pady=20)

        tk.Label(container, text="¿El control fue exitoso?").pack(pady=10)
        self.resultado_var = tk.StringVar(value="Exitoso")
        tk.OptionMenu(container, self.resultado_var, "Exitoso", "No Exitoso").pack()

        self.btn_registrar = tk.Button(container, text="Registrar Resultado", bg="#4CAF50", fg="white",
                                       command=self.validar_intento)
        self.btn_registrar.pack(pady=20)

        self.btn_excel = tk.Button(container, text="Informe Excel", state="disabled", command=self.exportar_informe)
        self.btn_excel.pack()

        tk.Button(container, text="Volver al Menú Principal", bg="#FFCCCB", width=20,command=self.regresar_inicio).pack(side="bottom", pady=40)

    def regresar_inicio(self):
        """Limpia los datos del paciente actual y vuelve al inicio"""
        self.intento_actual = 1
        self.historial_intentos = {}
        self.datos_paciente = {}
        self.create_main_menu()

    def validar_intento(self):
        resultado = self.resultado_var.get()
        tecnica = self.entries["Tecnica"].get()

        # Guardar en historial para el Excel
        self.historial_intentos[self.intento_actual] = {"paquete": self.paquete_actual_str, "resultado": resultado}

        if resultado == "Exitoso":
            messagebox.showinfo("Éxito", "Control validado correctamente.")
            self.btn_registrar.config(state="disabled")
            self.btn_excel.config(state="normal", bg="#0078D7", fg="white")
        else:
            # Lógica de fallo
            es_vmat_simple = False
            try:
                if tecnica in ["IMRT", "VMAT"] and not (
                        float(self.datos_paciente["MCSmin"]) < 0.5 or float(self.datos_paciente["SASmax"]) > 0.5 or int(
                        self.datos_paciente["Fractions"]) > 3):
                    es_vmat_simple = True
            except:
                pass

            if es_vmat_simple or self.intento_actual >= 2:
                messagebox.showerror("CRÍTICO", "EL CONTROL HA FALLADO.\n\nSE DEBE REHACER EL PLAN DE TRATAMIENTO.")
                self.btn_registrar.config(state="disabled")
                # Permitimos exportar el fallo para que quede registro de que se debe rehacer
                self.btn_excel.config(state="normal", bg="#D9534F", fg="white")
            else:
                self.intento_actual += 1
                messagebox.showwarning("Fallo", "Primer intento fallido. Pase al segundo control de QA.")
                self.ejecutar_arbol_decision()

    def obtener_costo_acumulado(self):
        """Busca y suma los costos de todas las técnicas usadas en los intentos"""
        ruta_costos = "costos.xlsx"
        if not os.path.exists(ruta_costos):
            return 0

        try:
            # Leemos el archivo. Usamos iloc para asegurar las columnas 0 (A) y 10 (K)
            df_costos = pd.read_excel(ruta_costos)

            # Creamos un diccionario { "Nombre Tecnica": Costo }
            # .strip() elimina espacios accidentales para que coincida con el paquete
            mapa_precios = pd.Series(
                df_costos.iloc[:, 10].values,
                index=df_costos.iloc[:, 0].astype(str).str.strip()
            ).to_dict()

            costo_total = 0
            for intento in self.historial_intentos.values():
                paquete = intento.get("paquete", "")
                if paquete and paquete != "-":
                    # Separamos el paquete "Plancheck + LogFile" -> ["Plancheck", "LogFile"]
                    tecnicas_individuales = [t.strip() for t in paquete.split("+")]
                    for t in tecnicas_individuales:
                        # Sumamos el costo si existe en el Excel, de lo contrario sumamos 0
                        costo_total += mapa_precios.get(t, 0)

            return round(float(costo_total),2)
        except Exception as e:
            print(f"Error al leer costos: {e}")
            return 0.00

    def exportar_informe(self):
        if not self.ruta_informe:
            messagebox.showwarning("Atención", "No hay una ruta definida para el registro.")
            self.seleccionar_registro_existente()
            if not self.ruta_informe: return

        # 1. Calculamos el costo asociado sumando todos los intentos
        costo_asociado = self.obtener_costo_acumulado()

        # 2. Construimos la fila para el Excel
        fila = {
            "Fecha": datetime.now().strftime("%d/%m/%Y"),
            "ID": self.datos_paciente.get("ID", "-"),
            "Paciente": self.datos_paciente.get("Nombre", "-"),
            "Técnica RT": self.entries["Tecnica"].get(),
            "MCS Min": self.datos_paciente.get("MCSmin"),
            "SAS Max": self.datos_paciente.get("SASmax")
        }

        # Registrar intentos (Máximo 2 según tu nueva lógica)
        for i in [1, 2]:
            info = self.historial_intentos.get(i, {"paquete": "-", "resultado": "-"})
            fila[f"QA Intento {i}"] = info["paquete"]
            fila[f"Resultado {i}"] = info["resultado"]

        # 3. Agregamos la columna de costo al final
        fila["Costo asociado"] = costo_asociado

        try:
            if os.path.exists(self.ruta_informe):
                df_existente = pd.read_excel(self.ruta_informe)
                df_final = pd.concat([df_existente, pd.DataFrame([fila])], ignore_index=True, sort=False)
            else:
                df_final = pd.DataFrame([fila])

            df_final.to_excel(self.ruta_informe, index=False)

            # Aplicamos el formato visual (colores y anchos)
            self.aplicar_formato_excel(self.ruta_informe)

            messagebox.showinfo("Éxito", f"Informe actualizado")
            self.btn_excel.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el informe: {e}")

    def aplicar_formato_excel(self, ruta):
        wb = load_workbook(ruta)
        ws = wb.active

        # Estilos previos
        relleno_header = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        fuente_header = Font(bold=True)
        borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # 1. Aplicar estilo general a cabeceras y encontrar la columna de costo
        col_costo_idx = None
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment, cell.border = relleno_header, fuente_header, Alignment(
                horizontal="center"), borde
            if cell.value == "Costo asociado":
                col_costo_idx = cell.column  # Guardamos el índice de la columna

        # 2. Aplicar estilos a las filas de datos
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border, cell.alignment = borde, Alignment(horizontal="left")

                # SI es la celda de la columna de costo, aplicamos el formato $0,00
                if cell.column == col_costo_idx:
                    # El formato '"$"#,##0.00' mostrará el punto o coma según la región de tu Windows/Excel
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")  # Alineamos moneda a la derecha

        # 3. Ajuste automático de ancho de columnas
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        wb.save(ruta)

if __name__ == "__main__":
    root = tk.Tk();
    app = RadioRiskApp(root);
    root.mainloop()