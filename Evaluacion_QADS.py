import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


class RadioRiskApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Evaluación de Riesgo en Radioterapia")
        self.ancho_fijo = 480
        self.alto_fijo = 600
        self.root.geometry(f"{self.ancho_fijo}x{self.alto_fijo}")

        # --- LÓGICA DE RUTAS PARA EJECUTABLE ---
        if getattr(sys, 'frozen', False):
            self.base_dir = os.path.dirname(sys.executable)
        else:
            self.base_dir = os.path.dirname(os.path.abspath(__file__))

        self.archivo_config = os.path.join(self.base_dir, "config_ruta.txt")
        self.archivo_umbrales = os.path.join(self.base_dir, "umbrales.txt")
        self.ruta_costos = os.path.join(self.base_dir, "costos.xlsx")

        # Intentar cargar ruta guardada, si no, usar la local por defecto
        self.ruta_informe = self.cargar_ruta_persistente()

        self.datos_paciente = {}
        self.entries = {}
        self.intento_actual = 1
        self.historial_intentos = {}

        # Umbrales (con carga segura)
        self.u_mcs, self.u_sas, self.u_dpf = 0.5, 0.5, 300
        self.u_mcs_min, self.u_sas_max, self.u_pmu = 0.5, 0.5, 1000
        self.cargar_umbrales()
        self.create_main_menu()

    def cargar_umbrales(self):
        if os.path.exists(self.archivo_umbrales):
            try:
                with open(self.archivo_umbrales, "r") as f:
                    lineas = f.readlines()
                    self.u_mcs = float(lineas[0].strip())
                    self.u_sas = float(lineas[1].strip())
                    self.u_dpf = float(lineas[2].strip())
                    self.u_mcs_min = float(lineas[3].strip())
                    self.u_sas_max = float(lineas[4].strip())
                    self.u_pmu = int(lineas[5].strip())
            except:
                pass

    def guardar_umbrales_archivo(self, mcs, sas, dpf, mcs_min, sas_max, pmu):
        with open(self.archivo_umbrales, "w") as f:
            f.write(f"{mcs}\n{sas}\n{dpf}\n{mcs_min}\n{sas_max}\n{pmu}")
        self.u_mcs, self.u_sas, self.u_dpf, self.u_mcs_min, self.u_sas_max, self.u_pmu = mcs, sas, dpf, mcs_min, sas_max, pmu

    def validar_y_guardar_umbrales(self):
        try:
            mcs = float(self.tmp_mcs.get().replace(',', '.'))
            sas = float(self.tmp_sas.get().replace(',', '.'))
            dpf = float(self.tmp_dpf.get().replace(',', '.'))
            mcs_min = float(self.tmp_mcs_min.get().replace(',', '.'))
            sas_max = float(self.tmp_sas_max.get().replace(',', '.'))
            pmu = int(self.tmp_pmu.get())

            self.guardar_umbrales_archivo(mcs, sas, dpf, mcs_min, sas_max, pmu)
            self.popup_silencioso("Éxito", "Umbrales actualizados correctamente.")
            self.create_config_menu()
        except ValueError:
            messagebox.showerror("Error", "Por favor, ingrese solo números válidos.")

    def cargar_ruta_persistente(self):
        # 1. Intentar leer desde config_ruta.txt
        if os.path.exists(self.archivo_config):
            with open(self.archivo_config, "r") as f:
                ruta = f.read().strip()
                if os.path.exists(ruta):
                    return ruta

        # 2. Si no hay config o la ruta no existe, buscar el archivo predeterminado en la carpeta
        ruta_defecto = os.path.join(self.base_dir, "Registro_Historico_2026.xlsx")
        if os.path.exists(ruta_defecto):
            return ruta_defecto

        return None

    def guardar_ruta_persistente(self, ruta):
        with open(self.archivo_config, "w") as f:
            f.write(ruta)
        self.ruta_informe = ruta

    def create_main_menu(self):
        for widget in self.root.winfo_children(): widget.destroy()
        menu_container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        menu_container.place(relx=0.5, rely=0.5, anchor="center")
        menu_container.pack_propagate(False)

        tk.Label(menu_container, text="Evaluación de Riesgo QADS", font=("Arial", 18, "bold")).pack(pady=40)
        tk.Button(menu_container, text="Cargar Paciente", width=25, height=2, bg="#E1E1E1", command=self.cargar_archivo,
                  font=("Arial", 10)).pack(pady=10)
        tk.Button(menu_container, text="Configuración", width=25, height=2, bg="#E1E1E1",
                  command=self.create_config_menu, font=("Arial", 10)).pack(pady=10)
        tk.Button(menu_container, text="Abrir Registros", width=25, height=2, bg="#E1E1E1",
                  command=self.abrir_excel_registros, font=("Arial", 10)).pack(pady=10)

        txt_ruta = os.path.basename(self.ruta_informe) if self.ruta_informe else "No configurada"
        tk.Label(menu_container, text=f"Ruta actual: {txt_ruta}", font=("Arial", 8, "italic"), fg="gray").pack(
            side="bottom", pady=20)

    def create_config_menu(self):
        for widget in self.root.winfo_children(): widget.destroy()
        c = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        c.place(relx=0.5, rely=0.5, anchor="center")
        c.pack_propagate(False)

        tk.Label(c, text="Panel de Configuración", font=("Arial", 16, "bold")).pack(pady=20)
        tk.Button(c, text="Seleccionar Registro Existente", width=30, height=2, bg="#E1E1E1",
                  command=self.seleccionar_registro_existente).pack(pady=10)
        tk.Button(c, text="Crear Nuevo Registro", width=30, height=2, bg="#E1E1E1",
                  command=self.crear_nuevo_registro).pack(pady=5)
        tk.Button(c, text="Configurar Umbrales", width=30, height=2, bg="#E1E1E1",
                  command=self.create_thresholds_menu).pack(pady=5)
        tk.Button(c, text="Configurar Costos", width=30, height=2, bg="#E1E1E1", command=self.abrir_excel_costos).pack(
            pady=5)
        tk.Button(c, text="Volver al Menú Principal", bg="#FFCCCB", command=self.create_main_menu).pack(side="bottom",
                                                                                                        pady=30)

    def create_thresholds_menu(self):
        for widget in self.root.winfo_children(): widget.destroy()
        container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        tk.Label(container, text="Configurar Umbrales de Complejidad", font=("Arial", 14, "bold")).pack(pady=(30, 20))
        frame_form = tk.Frame(container)
        frame_form.pack(expand=True)

        self.tmp_mcs = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_mcs.insert(0, str(self.u_mcs))
        self.tmp_sas = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_sas.insert(0, str(self.u_sas))
        self.tmp_dpf = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_dpf.insert(0, str(self.u_dpf))
        self.tmp_mcs_min = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_mcs_min.insert(0, str(self.u_mcs_min))
        self.tmp_sas_max = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_sas_max.insert(0, str(self.u_sas_max))
        self.tmp_pmu = tk.Entry(frame_form, justify="center", width=12)
        self.tmp_pmu.insert(0, str(self.u_pmu))

        campos = [("MCS Mínimo (Prom.):", self.tmp_mcs), ("SAS Máximo (Prom.):", self.tmp_sas),
                  ("dpf Mínimo:", self.tmp_dpf), ("MCS Mínimo:", self.tmp_mcs_min),
                  ("SAS Máximo:", self.tmp_sas_max), ("PMU:", self.tmp_pmu)]

        for i, (texto, entry) in enumerate(campos):
            tk.Label(frame_form, text=texto, font=("Arial", 10, "bold"), anchor="e", width=18).grid(row=i, column=0,
                                                                                                    pady=10, padx=10,
                                                                                                    sticky="e")
            entry.grid(row=i, column=1, pady=15, padx=15, sticky="w")

        btn_frame = tk.Frame(container)
        btn_frame.pack(pady=40)
        tk.Button(btn_frame, text="Guardar Cambios", bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), width=20,
                  height=2, command=self.validar_y_guardar_umbrales).pack(pady=5)
        tk.Button(btn_frame, text="Volver", font=("Arial", 10), width=15, command=self.create_config_menu).pack(pady=5)

    def seleccionar_registro_existente(self):
        ruta = filedialog.askopenfilename(title="Seleccione el archivo de Registro Histórico",
                                          filetypes=[("Excel files", "*.xlsx")])
        if ruta:
            self.guardar_ruta_persistente(ruta)
            self.popup_silencioso("Éxito", f"Se ha vinculado el archivo:\n{os.path.basename(ruta)}")
            self.create_main_menu()

    def crear_nuevo_registro(self):
        nueva_ruta = filedialog.asksaveasfilename(title="Nuevo Registro Histórico", defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx")],
                                                  initialfile="Registro_Historico_2026.xlsx")
        if nueva_ruta:
            self.guardar_ruta_persistente(nueva_ruta)
            self.popup_silencioso("Éxito", f"Nuevo registro configurado en:\n{os.path.basename(nueva_ruta)}")
            self.create_main_menu()

    def abrir_excel_costos(self):
        if os.path.exists(self.ruta_costos):
            try:
                os.startfile(self.ruta_costos)
                self.popup_silencioso("Costos", "Abriendo archivo de costos...")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir: {e}")
        else:
            messagebox.showerror("Error", f"No se encontró el archivo 'costos.xlsx' en:\n{self.base_dir}")

    def abrir_excel_registros(self):
        if self.ruta_informe and os.path.exists(self.ruta_informe):
            try:
                os.startfile(self.ruta_informe)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir: {e}")
        else:
            messagebox.showwarning("Atención",
                                   "No hay un archivo de registro creado todavía o no se ha seleccionado ninguno.")

    def cargar_archivo(self):
        filepath = filedialog.askopenfilename(title="Seleccionar reporte", filetypes=[("Excel files", "*.xlsx *.xls")])
        if filepath:
            try:
                self.intento_actual = 1
                self.historial_intentos = {}
                self.extraer_datos(filepath)
                self.mostrar_detalles_paciente()
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer el archivo: {e}")

    def extraer_datos(self, path):
        df = pd.read_excel(path, header=None)

        def buscar_valor(etiqueta):
            for i, row in df.iterrows():
                for j, cell in enumerate(row):
                    if str(cell).strip().upper() == etiqueta.upper():
                        return str(df.iloc[i, j + 1]).strip()
            return "-"

        mcs_values, sas_values = [], []
        en_beam_metrics = False
        for i, row in df.iterrows():
            if str(row[0]).strip().upper() == "BEAM METRICS":
                en_beam_metrics = True
                continue
            if en_beam_metrics:
                try:
                    metrica, valor_str = str(row[2]).strip(), str(row[3]).replace(',', '.')
                    if metrica == "MCS":
                        mcs_values.append(float(valor_str))
                    elif metrica == "SAS":
                        sas_values.append(float(valor_str))
                except:
                    continue

        self.datos_paciente = {
            "Plan": buscar_valor("PLAN NAME"),
            "Nombre": buscar_valor("PATIENT NAME"),
            "ID": buscar_valor("PATIENT ID"),
            "Sexo": buscar_valor("PATIENT SEX"),
            "Fractions": buscar_valor("FRACTIONS"),
            "MCS": buscar_valor("MCS"),
            "SAS": buscar_valor("SAS"),
            "PMU": buscar_valor("PMU"),
            "MCSmin": str(min(mcs_values)) if mcs_values else "-",
            "SASmax": str(max(sas_values)) if sas_values else "-",
            "dpf": buscar_valor("DOSE/FRACTION [cGy]")
        }

    def actualizar_checkbox_ca(self, *args):
        region = self.entries["Region"].get()
        regiones_con_ca = ["COLON/RECTO", "PULMON", "CERVIX/UTERO", "CYC"]
        self.entries["CA"].set(region in regiones_con_ca)

    def mostrar_detalles_paciente(self):
        for widget in self.root.winfo_children(): widget.destroy()
        container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        tk.Label(container, text="Información del Paciente", font=("Arial", 14, "bold")).pack(pady=10)
        frame_info = tk.LabelFrame(container, text=" Datos Extraídos ", padx=15, pady=10)
        frame_info.pack(padx=10, fill="both", expand=True)

        op_sexo, op_tecnica = ["M", "F", "-"], ["3D", "IMRT", "VMAT", "SRS", "SBRT", "FIF"]
        op_anatomica = ["MAMA", "COLON/RECTO", "PULMON", "PROSTATA", "CERVIX/UTERO", "ESOFAGO", "CYC", "PANCREAS",
                        "VEJIGA", "ENCEFALO/SNC", "MIEMBROS", "OTROS"]

        plan_name = self.datos_paciente.get("Plan", "").upper()
        tecnica_def = next((t for t in op_tecnica if t in plan_name), "3D")

        self.entries["Sexo"] = tk.StringVar(value=self.datos_paciente.get("Sexo", "-"))
        self.entries["Region"] = tk.StringVar(value="OTROS")
        self.entries["Tecnica"] = tk.StringVar(value=tecnica_def)
        self.entries["CA"] = tk.BooleanVar()
        self.entries["PPed"] = tk.BooleanVar()
        self.entries["Region"].trace_add("write", self.actualizar_checkbox_ca)
        self.actualizar_checkbox_ca()

        campos = [("ID Paciente", "ID"), ("Nombre", "Nombre"), ("Plan", "Plan"), ("MCS Prom.", "MCS"),
                  ("SAS Prom.", "SAS"), ("PMU", "PMU"), ("MCS Mínimo", "MCSmin"), ("SAS Máximo", "SASmax"),
                  ("Fracciones", "Fractions"), ("dpf [cGy]", "dpf")]
        for label, key in campos:
            row = tk.Frame(frame_info);
            row.pack(fill="x", pady=1)
            tk.Label(row, text=label, width=15, anchor="w", font=("Arial", 9)).pack(side="left")
            e = tk.Entry(row, font=("Arial", 9));
            e.insert(0, self.datos_paciente.get(key, "-"))
            e.config(state="readonly");
            e.pack(side="right", fill="x", expand=True)

        for lab, key, ops in [("Sexo", "Sexo", op_sexo), ("Región", "Region", op_anatomica),
                              ("Técnica", "Tecnica", op_tecnica)]:
            row = tk.Frame(frame_info);
            row.pack(fill="x", pady=2)
            tk.Label(row, text=lab, width=15, anchor="w", font=("Arial", 9, "bold")).pack(side="left")
            tk.OptionMenu(row, self.entries[key], *ops).pack(side="right", fill="x", expand=True)

        tk.Checkbutton(frame_info, text="Cambios Anatómicos", variable=self.entries["CA"]).pack(anchor="w")
        tk.Checkbutton(frame_info, text="Paciente Pediátrico", variable=self.entries["PPed"]).pack(anchor="w")
        tk.Button(container, text="Calcular Método QA", bg="#0078D7", fg="white", font=("Arial", 10, "bold"),
                  command=self.ejecutar_arbol_decision).pack(pady=10)
        tk.Button(container, text="Volver", command=self.create_main_menu).pack()

    def es_plan_complejo(self):
        if self.entries["Tecnica"].get() not in ["IMRT", "VMAT"]: return False

        def limpiar(v, d):
            if v in [None, "-", ""]: return d
            try:
                return float(str(v).replace(',', '.'))
            except:
                return d

        condiciones = [
            limpiar(self.datos_paciente.get("MCSmin"), 1.0) < self.u_mcs_min,
            limpiar(self.datos_paciente.get("SASmax"), 0.0) > self.u_sas_max,
            float(limpiar(self.datos_paciente.get("dpf"), 0.0)) > self.u_dpf,
            limpiar(self.datos_paciente.get("MCS"), 1.0) < self.u_mcs,
            limpiar(self.datos_paciente.get("SAS"), 0.0) > self.u_sas,
            limpiar(self.datos_paciente.get("PMU"), 0) > self.u_pmu
        ]
        return any(condiciones)

    def obtener_paquete_qa(self):
        tecnica = self.entries["Tecnica"].get()
        ca_ped = self.entries["CA"].get() or self.entries["PPed"].get()
        complejo = self.es_plan_complejo()

        if tecnica in ["3D", "FIF"]:
            res = "Plancheck + Calculo independiente + LogFile" if self.intento_actual == 1 else "Portal Dosimetry"
            return res + " + Transit-EPID" if ca_ped and self.intento_actual == 1 else res
        elif tecnica in ["SRS", "SBRT"]:
            res = "Plancheck + Portal Dosimetry" if self.intento_actual == 1 else "Stereophan + Gafchromic/CI"
            return res + " + Transit-EPID" if ca_ped and self.intento_actual == 1 else res
        elif tecnica in ["IMRT", "VMAT"]:
            if not complejo:
                res = "Plancheck + Calculo independiente + LogFile" if self.intento_actual == 1 else "Portal Dosimetry"
            else:
                res = "Plancheck + Calculo independiente + LogFile + Portal Dosimetry" if self.intento_actual == 1 else "ArcCheck + 3DVH"
            return res + " + Transit-EPID" if ca_ped and self.intento_actual == 1 else res
        return "Indefinido"

    def ejecutar_arbol_decision(self):
        for widget in self.root.winfo_children(): widget.destroy()
        container = tk.Frame(self.root, width=self.ancho_fijo, height=self.alto_fijo)
        container.place(relx=0.5, rely=0.5, anchor="center");
        container.pack_propagate(False)

        paquete = self.obtener_paquete_qa()
        self.paquete_actual_str = paquete
        tk.Label(container, text="EVALUACIÓN QADS", font=("Arial", 14, "bold")).pack(pady=20)
        tk.Label(container, text=f"Intento N° {self.intento_actual}", font=("Arial", 10, "italic")).pack()
        tk.Label(container, text=paquete, font=("Arial", 11, "bold"), fg="#004080", wraplength=400).pack(pady=20)
        tk.Label(container, text="¿El control fue exitoso?").pack(pady=10)
        self.resultado_var = tk.StringVar(value="Exitoso")
        tk.OptionMenu(container, self.resultado_var, "Exitoso", "No Exitoso").pack()

        self.btn_registrar = tk.Button(container, text="Registrar Resultado", bg="#4CAF50", fg="white",
                                       command=self.validar_intento)
        self.btn_registrar.pack(pady=20)
        self.btn_excel = tk.Button(container, text="Informe Excel", state="disabled", command=self.exportar_informe)
        self.btn_excel.pack()
        tk.Button(container, text="Volver al Inicio", bg="#FFCCCB", width=20, command=self.regresar_inicio).pack(
            side="bottom", pady=40)

    def regresar_inicio(self):
        self.intento_actual = 1;
        self.historial_intentos = {};
        self.datos_paciente = {}
        self.create_main_menu()

    def validar_intento(self):
        resultado = self.resultado_var.get()
        tecnica = self.entries["Tecnica"].get()
        self.historial_intentos[self.intento_actual] = {"paquete": self.paquete_actual_str, "resultado": resultado}

        if resultado == "Exitoso":
            self.popup_silencioso("Éxito", "Control validado correctamente.")
            self.btn_registrar.config(state="disabled")
            self.btn_excel.config(state="normal", bg="#0078D7", fg="white")
        else:
            es_complejo = self.es_plan_complejo()
            if (tecnica in ["IMRT", "VMAT"] and not es_complejo) or self.intento_actual >= 2 or tecnica in ["3D",
                                                                                                            "FIF"]:
                messagebox.showerror("CRÍTICO", "EL CONTROL HA FALLADO.\n\nSE DEBE REHACER EL PLAN DE TRATAMIENTO.")
                self.btn_registrar.config(state="disabled")
                self.btn_excel.config(state="normal", bg="#D9534F", fg="white")
            else:
                self.intento_actual += 1
                messagebox.showerror("Fallo", f"Intento {self.intento_actual - 1} fallido. Pase al siguiente escalón.")
                self.ejecutar_arbol_decision()

    def obtener_costo_acumulado(self):
        if not os.path.exists(self.ruta_costos): return 0.0
        try:
            df_costos = pd.read_excel(self.ruta_costos)
            mapa_precios = pd.Series(df_costos.iloc[:, 10].values,
                                     index=df_costos.iloc[:, 0].astype(str).str.strip()).to_dict()
            costo_total = 0
            for intento in self.historial_intentos.values():
                for t in [x.strip() for x in intento.get("paquete", "").split("+")]:
                    costo_total += mapa_precios.get(t, 0)
            return round(float(costo_total), 2)
        except:
            return 0.0

    def exportar_informe(self):
        if not self.ruta_informe:
            messagebox.showwarning("Atención", "No hay una ruta definida.")
            self.seleccionar_registro_existente()
            if not self.ruta_informe: return

        fila = {
            "Fecha": datetime.now().strftime("%d/%m/%Y"),
            "ID": self.datos_paciente.get("ID", "-"),
            "Paciente": self.datos_paciente.get("Nombre", "-"),
            "Técnica RT": self.entries["Tecnica"].get(),
            "MCS Min": self.datos_paciente.get("MCSmin"),
            "SAS Max": self.datos_paciente.get("SASmax"),
            "Dosis/Frac": self.datos_paciente.get("dpf", "-")
        }
        for i in [1, 2]:
            info = self.historial_intentos.get(i, {"paquete": "-", "resultado": "-"})
            fila[f"QA Intento {i}"] = info["paquete"]
            fila[f"Resultado {i}"] = info["resultado"]
        fila["Costo asociado"] = self.obtener_costo_acumulado()

        try:
            if os.path.exists(self.ruta_informe):
                df_exist = pd.read_excel(self.ruta_informe)
                df_final = pd.concat([df_exist, pd.DataFrame([fila])], ignore_index=True, sort=False)
            else:
                df_final = pd.DataFrame([fila])

            df_final.to_excel(self.ruta_informe, index=False)
            self.aplicar_formato_excel(self.ruta_informe)
            self.popup_silencioso("Éxito", "Informe guardado correctamente.")
            self.btn_excel.config(state="disabled")
        except PermissionError:
            messagebox.showerror("Error", "El archivo Excel está abierto. Ciérrelo e intente de nuevo.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar: {e}")

    def aplicar_formato_excel(self, ruta):
        wb = load_workbook(ruta);
        ws = wb.active
        relleno = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                       bottom=Side(style='thin'))

        col_costo = None
        for cell in ws[1]:
            cell.fill, cell.font, cell.border = relleno, Font(bold=True), borde
            cell.alignment = Alignment(horizontal="center")
            if cell.value == "Costo asociado": col_costo = cell.column

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = borde
                if cell.column == col_costo:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        wb.save(ruta)

    def popup_silencioso(self, titulo, mensaje):
        vent = tk.Toplevel(self.root);
        vent.title(titulo);
        vent.geometry("300x150")
        vent.resizable(False, False);
        vent.transient(self.root);
        vent.grab_set()
        x = self.root.winfo_x() + 90;
        y = self.root.winfo_y() + 200
        vent.geometry(f"+{x}+{y}")
        tk.Label(vent, text=mensaje, wraplength=250, pady=20).pack()
        tk.Button(vent, text="Aceptar", width=10, command=vent.destroy, bg="#E1E1E1").pack(pady=10)


if __name__ == "__main__":
    root = tk.Tk();
    app = RadioRiskApp(root);
    root.mainloop()